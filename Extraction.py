import numpy as np
import cv2
import pywt
import glob
from PIL import Image
import copy
import openpyxl
import math
from matplotlib import pyplot as plt
import pickle

def mse(I1, I2):
    err = np.square(np.subtract(np.double(I1), np.double(I2))).mean()
    return err

def psnr(img1, img2):
    mseval = mse(img1, img2)
    if mse == 0:
        return 100
    PIXEL_MAX = 255.0
    return 20 * math.log10(PIXEL_MAX / math.sqrt(mseval))

def halt():
    k = 0
    while k != 48:
        k = cv2.waitKey(100)
    cv2.destroyAllWindows()

def Save_Image(Location, FName2, image_array):
    fn = FName2
    LenFn=len(fn)
    fn2=fn[0:LenFn-4]+'.png'
    img = Image.fromarray(image_array)
    #Floc = Location + fn
    Floc2= Location + fn2
    #img.save(Floc)  # save image at location
    img.save(Floc2)
    print('Image Saved at: ' + Floc2)
    return

def DFT_Decoding(watermarkedImage,coverImageFFT,alpha,watermarkImage2):
    ## Extraction process
    watermarkedImageR = watermarkedImage / 255  # convert to 0,1 range
    watermarkedImageRFFT = np.fft.fftshift(np.fft.fft2(watermarkedImageR))  # apply FFT
    watermarkRFFT = (watermarkedImageRFFT - coverImageFFT) / alpha  # extract watermark FFT
    watermarkR = np.fft.ifft2(np.fft.ifftshift(watermarkRFFT))  # apply inverse FFT
    watermarkR = np.uint8(255 * watermarkR)  # convert to 0,255 range
    # cv2.imshow('Extracted watermark',watermarkR)

    ## calculate MSE
    #watermarkImage2 = np.uint8(255 * watermarkImage2)
    watermarkImage2=np.uint8(watermarkImage2)

    MSE = mse(watermarkImage2, watermarkR)
    print('MSE (watermark, extracted watermark)=' + str(MSE))
    PSNR = psnr(watermarkImage2, watermarkR)
    print('PSNR (watermark, extracted watermark)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    XLS_path = "watermarking decoding results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    print('Row No. : '+str(CountROW+1))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DFT'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save(XLS_path)

    # save extracted watermark image

    #img = Image.fromarray(watermarkR)
    #Floc = './watermark_recovered/DFT/' + fn  # set file path+file name
    #img.save(Floc)  # save image at location
    #print('Extracted Watermark Image saved at: ' + Floc)
    Save_Image('./watermark_recovered/dft/', FName, watermarkR)

    return
# option 3
def DFT(watermarkImage, FName):
    alpha = 0.05
    print("--- Decoding Starts ---")

    # read watermarked image
    NFName=len(FName)
    WFileName='./watermarked/dft/'+FName[0:(NFName-4)]+'.png'
    watermarkedImage = cv2.imread(WFileName, 0)
    H, W = np.shape(watermarkedImage)

    # resize watermark image
    watermarkImage2 = cv2.resize(watermarkImage, (W, H))

    # load key values
    VarFileName = './key/dft/' + 'key_' + FName[0:len(FName)-4] + '.jpg.pkl'
    FileObj = open(VarFileName, 'rb')
    coverImageFFT = pickle.load(FileObj)

    DFT_Decoding(watermarkedImage, coverImageFFT,alpha,watermarkImage2)
    return

def DWT_Decoding(watermarkedImage, watermarkImage, alpha, cA):
    ## DECODING DWT WITHOUT NOISE
    # print("******* WITHOUT NOISE ---------")
    #extracted = DWT_Decoding(watermarkedImage, alpha, cA)
    #Extraction
    watermarkedImage=watermarkedImage/255
    coeffWM = pywt.dwt2(watermarkedImage, 'haar')
    hA, (hH, hV, hD) = coeffWM

    extracted = (hA-cA)/alpha
    extracted = np.uint8(255*extracted)

    # cv2.imshow('Extracted Watermark (without Noise)',extracted)
    print("Watermark Extracted...")
    MSE = mse(watermarkImage, extracted)
    print('MSE (watermark, extracted watermark)=' + str(MSE))
    PSNR = psnr(watermarkImage, extracted)
    print('PSNR (watermark, extracted watermark)=' + str(PSNR))
    
    # save recovered watermark image
    Save_Image('./watermark_recovered/DWT/', FName, extracted)

    ## UPDATE EXCEL DATA
    XLS_path = "watermarking decoding results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    print('Row No. : '+str(CountROW+1))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DWT'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save(XLS_path)
    
    return ()
# option 1
def DWT(watermarkImage, FName):
    print("--- DWT Watermarking ---")
    alpha = 0.1
#    coverImage = cv2.resize(coverImage, (300, 300))
    watermarkImage = cv2.resize(watermarkImage, (150, 150))

    NFName=len(FName)
    WFileName='./watermarked/dwt/'+FName[0:(NFName-4)]+'.png'
    watermarkedImage = cv2.imread(WFileName, 0)

    # load key values
    VarFileName = './key/dwt/' + 'key_' + FName[0:len(FName)-4] + '.jpg.pkl'
    FileObj = open(VarFileName, 'rb')
    cA = pickle.load(FileObj)

    DWT_Decoding(watermarkedImage, watermarkImage, alpha, cA)

    return watermarkedImage

def DCT_Decoding(Imask, Iw, Zp, DctOrg, alpha):
    print("--- Decoding Starts ---")
    # Extraction process
    [Hw, Ww] = np.shape(Imask)
    NWp = len(DctOrg)
    ## Lower DCT Coefficient locations of 8x8 block
    RC = [[0, 0], [1, 0], [2, 0], [3, 0], [4, 0], [5, 0], [6, 0],
          [0, 1], [1, 1], [2, 1], [3, 1], [4, 1], [5, 1],
          [0, 2], [1, 2], [2, 2], [3, 2], [4, 2],
          [0, 3], [1, 3], [2, 3], [3, 3],
          [0, 4], [1, 4], [2, 4],
          [0, 5], [1, 5],
          [0, 6]]
    NALoc = RC.__len__()
    [H, W]=np.shape(Iw)
    Iw = np.uint8(Iw)
    Wc = 0
    # Im=I
    Imr = np.zeros((NWp, 1))
    DctVal = np.zeros((NALoc, 1))
    for i in range(0, H - 8, 8):
        for j in range(0, W - 8, 8):
            Block = Iw[i:i + 8, j: j + 8]  # crop image block 8 x8
            DctBlock = cv2.dct(np.double(Block))  # convert to dct

            for k in range(0, NALoc):
                r = RC[k][0]
                c = RC[k][1]
                DctVal[k] = DctBlock[r, c]  # find lower dc component end

            Wc2 = Wc + NALoc  # end location
            #print('i=' + str(i) + ', ' + 'j=' + str(j) + ', Wc=' + str(Wc) + ', Wc2=' + str(Wc2))

            # DctOrg = np.zeros((NWp, 1))
            DctOrgt = DctOrg[Wc:Wc2]  # copy original data
            DCTRec = abs(DctVal - DctOrgt) / alpha
            Imr[Wc: Wc2] = DCTRec
            if (Wc > (NWp - NALoc - 1)):
                break
            else:
                Wc = Wc + NALoc
        if (Wc > (NWp - NALoc - 1)):
            break
    ## reshape and display
    L = Imr.__len__()
    Imru = Imr[0:L - Zp]
    Imru = np.resize(Imru, (Hw, Ww))
    Imru = np.double(255 * Imru)
    Imru = cv2.threshold(Imru, 127, 255, cv2.THRESH_BINARY)[1]

    Imru = Imru / 255
    # cv2.imshow('Extracted Watermark: '+FName,Imru)

    # calculate MSE
    Imask2 = copy.deepcopy(Imask)
    Imru2 = copy.deepcopy(Imru)
    Imask2 = np.uint8(255 * Imask)
    Imru2 = np.uint8(255 * Imru)
    #img = Image.fromarray((Imru2))
    #img.show()

    MSE = mse(Imask2, Imru2)
    print('MSE (watermark , extracted watermarked)=' + str(MSE))
    PSNR = psnr(Imask2, Imru2)
    print('PSNR (watermark , extracted watermarked)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    XLS_path = "watermarking decoding results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    print('Row No. : '+str(CountROW+1))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DCT'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save(XLS_path)

    Imru = np.uint8(255 * Imru)

    # save recovered watermark image
    Save_Image('./watermark_recovered/dct/', FName, Imru)
    return
# option 2
def DCT(Imask2, FName):
    print("--- DCT Watermarking ---")
    alpha = 2
    # read watermarked image
    NFName=len(FName)
    WFileName='./watermarked/dct/'+FName[0:(NFName-4)]+'.png'
    watermarkedImage = cv2.imread(WFileName, 0)

    # load key values
    VarFileName = './key/dct/' + 'key_' + FName[0:len(FName)-4] + '.jpg.pkl'
    FileObj = open(VarFileName, 'rb')
    [Imask, Zp, DctOrg] = pickle.load(FileObj)
    
    DCT_Decoding(Imask, watermarkedImage, Zp, DctOrg, alpha)
    return

def SVD_Decoding(watermarkedImage, U_SHL_w, V_SHL_w, Simg_temp, alpha, watermarkImage2):
    ##--------------- EXTRACTION PROCESS
    # Appyly SVD on watermarked image
    [H, W] = np.shape(watermarkedImage)
    m = H
    n = W
    if n < m:
        m = n
    else:
        n = m
        
    #[H, W] = np.shape(coverImage)
    #m = H
    #n = W
    #if n < m:
    #    m = n
    #else:
    #    n = m
    watermarkedImage = cv2.resize(watermarkedImage, (n, m))  # resize to size of cover image
    watermarkImage2 = cv2.resize(watermarkImage2, (n, m))
    Wimg, SWimg, VWimg = np.linalg.svd(watermarkedImage, full_matrices=1, compute_uv=1)

    # performing inverse SVD
    D_1 = np.dot((U_SHL_w * SWimg), V_SHL_w)
    # extracting watermark
    Watermark = np.abs(D_1 - Simg_temp) / alpha

    Watermark2 = 255 * Watermark
    Watermark2 = cv2.threshold(Watermark2, 127, 255, cv2.THRESH_BINARY)[1]
    Watermark2 = np.uint8(Watermark2)
    # cv2.imshow('Extracted Watermark :'+FName,Watermark2)

    MSE = mse(watermarkImage2, Watermark2)
    print('MSE (watermark, extracted watermark )= ' + str(MSE))
    PSNR = psnr(watermarkImage2, Watermark2)
    print('PSNR (watermark, extracted watermark )=' + str(PSNR))

    ## UPDATE EXCEL DATA
    XLS_path = "watermarking decoding results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    print('Row No. : '+str(CountROW+1))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'SVD'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save(XLS_path)

    [Hi, Wi] = np.shape(watermarkImage)
    Watermark2 = cv2.resize(watermarkImage2, (Wi, Hi))

    # save watermarked image
    Save_Image('./watermark_recovered/svd/', FName, Watermark2)
    return

# option 4
def SVD(watermarkImage, FName):
    print("--- SVD Watermarking ---")
    alpha = 10
    # read watermarked image
    NFName=len(FName)
    WFileName='./watermarked/svd/'+FName[0:(NFName-4)]+'.png'
    watermarkedImage = cv2.imread(WFileName, 0)
    H, W = np.shape(watermarkedImage)

    # resize watermark image
    watermarkImage2 = cv2.resize(watermarkImage, (W, H))
    # load key values
    VarFileName = './key/svd/' + 'key_' + FName[0:len(FName)-4] + '.jpg.pkl'
    FileObj = open(VarFileName, 'rb')
    [U_SHL_w, V_SHL_w, Simg_temp] = pickle.load(FileObj)

    SVD_Decoding(watermarkedImage, U_SHL_w, V_SHL_w, Simg_temp, alpha, watermarkImage2)
    return

def DWT_SVD_Decoding(watermarkedImage, watermarkImage, U_SHL_w, V_SHL_w, Simg_temp, alpha):
    ##--------------- EXTRACTION PROCESS
    # Apply DWT
    [Hr, Wr] = np.shape(watermarkedImage)
    mr = Hr
    nr = Wr
    if nr < mr:
        mr = nr
    else:
        nr = mr
    watermarkedImage = cv2.resize(watermarkedImage, (nr, mr))  # resize to size of cover image

    coeffC = pywt.dwt2(watermarkedImage, 'haar')
    A, (H, V, D) = coeffC
    # Apply SVD on A Band
    Wimg, SWimg, VWimg = np.linalg.svd(A, full_matrices=1, compute_uv=1)
    # performing inverse SVD
    D_1 = np.dot((U_SHL_w * SWimg), V_SHL_w)
    # extracting watermark
    Watermark = np.abs(D_1 - Simg_temp) / alpha

    Watermark2 = 255 * Watermark
    Watermark2 = cv2.threshold(Watermark2, 127, 255, cv2.THRESH_BINARY)[1]
    Watermark2 = np.uint8(Watermark2)
    [Hi, Wi] = np.shape(watermarkImage)
    Watermark2 = cv2.resize(Watermark2, (Wi, Hi))
    # plt.imshow(Watermark2,cmap='gray');plt.title('Extracted Watermark image');plt.show()

    MSE = mse(watermarkImage, Watermark2)
    print('MSE (cover image, watermarked image)=', MSE)
    PSNR = psnr(watermarkImage, Watermark2)
    print('PSNR (cover image, watermarked image)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    XLS_path = "watermarking decoding results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    print('Row No. : '+str(CountROW+1))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DWT-SVD'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save(XLS_path)

    # save extracted watermark image
    Save_Image('./watermark_recovered/dwtsvd/', FName, Watermark2)
    return

# option 5
def DWT_SVD(watermarkImage, FName):
    print("--- DWT SVD Watermarking ---")
    alpha = 10
    # read watermarked image
    NFName=len(FName)
    WFileName='./watermarked/dwtsvd/'+FName[0:(NFName-4)]+'.png'
    watermarkedImage = cv2.imread(WFileName, 0)
    H, W = np.shape(watermarkedImage)

    # resize watermark image
    watermarkImage2 = cv2.resize(watermarkImage, (W, H))
    # load key values
    VarFileName = './key/dwtsvd/' + 'key_' + FName[0:len(FName)-4] + '.jpg.pkl'
    FileObj = open(VarFileName, 'rb')
    [U_SHL_w, V_SHL_w, Simg_temp] = pickle.load(FileObj)

    DWT_SVD_Decoding(watermarkedImage, watermarkImage, U_SHL_w, V_SHL_w, Simg_temp, alpha)
    
    return watermarkedImage


def DWT_DCT_SVD_Decoding(watermarkedImage, watermarkImage, U_SHL_w, V_SHL_w, Simg_temp, alpha):
    ############## Apply Inverse Extraction process
    print('****Watermark Extraction******')
    coeffC = pywt.dwt2(watermarkedImage, 'haar')
    Da, (Dh, Dv, Dd) = coeffC
    print('dwt aaplied on image')
    [Ha, Wa] = np.shape(Da)

    # Generate DCT Block
    Bh = np.uint8(Ha / 8)  # total block can be processed in horizontal direction
    Bw = np.uint8(Wa / 8)  # total blocks can be processed in vertical direction

    # Extract DCT
    DMatD = np.zeros((Bh, Bw), np.float32)

    dblock = 7
    ih = 0
    iw = 0
    for i in range(0, Ha - 1, 8):
        iw = 0
        for j in range(0, Wa - 1, 8):
            Block = Da[i:i + 8, j:j + 8]
            DctBlock = cv2.dct(Block)
            DMatD[ih, iw] = DctBlock[dblock, dblock]
            iw = iw + 1
        ih = ih + 1
    print('DCT block generated')

    # Simg_temp=np.diag(Simg_temp);
    # Watermark=ISWD(DMatD,U_SHL_w,V_SHL_w,Simg_temp,alpha)
    [x, y] = np.shape(DMatD)
    Wimg, SWimg, VWimg = np.linalg.svd(DMatD, full_matrices=1, compute_uv=1)

    # apply inverse SVD, use U,V component of 2nd SVD
    # SWimg=np.diag(SWimg)
    # D_1=U_SHL_w * SWimg * V_SHL_w
    D_1 = np.dot((U_SHL_w * SWimg), V_SHL_w)
    # D_1=np.matmul(U_SHL_w,np.matmul(SWimg, V_SHL_w) )

    WatermarkR = (D_1 - Simg_temp) / alpha
    # Watermark=np.round(255*Watermark);
    WatermarkR = (255 * WatermarkR)
    (thresh, WatermarkR) = cv2.threshold(WatermarkR, 127, 255, cv2.THRESH_BINARY)

    print('Inverse SVD Performed, watermark image extracted')
    WatermarkR = np.uint8(WatermarkR)
    # cv2.namedWindow("Extracted Watermark Image: "+FName, cv2.WINDOW_NORMAL)
    # cv2.imshow('DWTDCTSVD_Extracted_Watermarked_'+FName,WatermarkR)

    # CALCULATE MSE
    MSE = mse(watermarkImage, WatermarkR)
    print('MSE (watermark, extracted watermark)=' + str(MSE))
    PSNR = psnr(watermarkImage, WatermarkR)
    print('PSNR (watermark, extracted watermark)=' + str(PSNR))

    ## UPDATE EXCEL DATA
    XLS_path = "watermarking decoding results.xlsx"
    XLS_Obj = openpyxl.load_workbook(XLS_path)
    XLS_Sheet_Obj = XLS_Obj.active
    CountROW = XLS_Sheet_Obj.max_row
    print('Row No. : '+str(CountROW+1))
    CELL1 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=1);
    CELL1.value = FName
    CELL2 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=2);
    CELL2.value = 'DWT-DCT-SVD'
    CELL3 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=3);
    CELL3.value = str(MSE)
    CELL4 = XLS_Sheet_Obj.cell(row=CountROW + 1, column=4);
    CELL4.value = str(PSNR)
    XLS_Obj.save(XLS_path)

    # save extracted watermark image
    Save_Image('./watermark_recovered/dwtdctsvd/', FName, WatermarkR)
    return

# Option 6
def DWT_DCT_SVD(watermarkImage, FName):
    print('--- DCT Watermarking ---')
    alpha = 0.1
    watermarkImage = cv2.resize(watermarkImage, (32, 32))
    # read watermarked image
    NFName=len(FName)
    WFileName='./watermarked/dwtdctsvd/'+FName[0:(NFName-4)]+'.png'
    watermarkedImage = cv2.imread(WFileName, 0)
    H, W = np.shape(watermarkedImage)

    # resize watermark image
    #watermarkImage2 = cv2.resize(watermarkImage, (W, H))
    # load key values
    VarFileName = './key/dwtdctsvd/' + 'key_' + FName[0:len(FName)-4] + '.jpg.pkl'
    FileObj = open(VarFileName, 'rb')
    [U_SHL_w, V_SHL_w, Simg_temp] = pickle.load(FileObj)

    DWT_DCT_SVD_Decoding(watermarkedImage, watermarkImage, U_SHL_w, V_SHL_w, Simg_temp, alpha)

    return

if __name__ == "__main__":

    cv2.destroyAllWindows()  # close all previos window
    watermarkImage = cv2.imread('watermarkImage4.JPG', 0)
    FuncName = ['DWT', 'DCT', 'DFT', 'SVD', 'DWT_SVD', 'DWT_DCT_SVD', 'DWT_DFT_SVD']
    options = {
        1: DWT,
        2: DCT,
        3: DFT,
        4: SVD,
        5: DWT_SVD,
        6: DWT_DCT_SVD,
    }

    val = input('What type of embedding you want to perform? \
                \n1.DWT\
                \n2.DCT\
                \n3.DFT\
                \n4.SVD\
                \n5.DWT-SVD\
                \n6.SVD-DCT-DWT\
                \n7.ALL_RUN\
                \nEnter your option: ')

    watermarking_function = options.get(int(val), None)
    sval = int(val)
    if sval < 7:
        if watermarking_function:
            if(val == '1'):
                fw = glob.glob('.\watermarked\dwt\*.png')  # search files with extension .jpg in same folder
            elif(val == '2'):
                fw = glob.glob('.\watermarked\dct\*.png')  # search files with extension .jpg in same folder
            elif(val == '3'):
                fw = glob.glob('.\watermarked\dft\*.png')  # search files with extension .jpg in same folder
            elif(val == '4'):
                fw = glob.glob('.\watermarked\svd\*.png')  # search files with extension .jpg in same folder
            elif(val == '5'):
                fw = glob.glob('.\watermarked\dwtsvd\*.png')  # search files with extension .jpg in same folder
            else:
                fw = glob.glob('.\watermarked\dwtdctsvd\*.png')  # search files with extension .jpg in same folder

            NFiles = len(fw)  # find total files found
            print('Files In folder')
            print(fw)  # print all file names
            
            for Fc in range(0, NFiles):  # run till all fies processed
                # read watermarked image
                stw = fw[Fc]  # select the file name
                if sval < 5:# cases 1 to 4
                    FName = stw[18:len(stw)]
                elif sval < 6:# case 5
                    FName = stw[21:len(stw)]
                else:# case for 6 and 7
                    FName = stw[24:len(stw)]

                print('\n**----------------------------------------')
                print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + stw)  # print file name
                watermarking_function(watermarkImage, FName)
            exit(1)
        else:
            print("Invalid Option")   
        
    else:# option 7 decode all methods
        print("All Test Run")
        print("1. -------------DWT WATERMARK EXTRACTION --------")
        fw = glob.glob('.\watermarked\dwt\*.png')
        NFiles = len(fw)
        print('Files In folder')
        print(fw)  # print all file names
        for Fc in range(0, NFiles):  # run till all fies processed
            stw = fw[Fc]  # select the file name
            FName = stw[18:len(stw)]
            FCoverImageName=FName[0:len(FName)-4]+'.jpg'
            coverImage = cv2.imread('./coverimages/'+FCoverImageName, 0)
            print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + stw)
            DWT(watermarkImage, FName)
            
        print("\n\n2. -------------DCT WATERMARK EXTRACTION --------")
        fw = glob.glob('.\watermarked\dct\*.png')
        NFiles = len(fw)
        print('Files In folder')
        print(fw)  # print all file names
        for Fc in range(0, NFiles):  # run till all fies processed
            stw = fw[Fc]  # select the file name
            FName = stw[18:len(stw)]
            FCoverImageName=FName[0:len(FName)-4]+'.jpg'
            coverImage = cv2.imread('./coverimages/'+FCoverImageName, 0)
            print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + stw)
            DCT(watermarkImage, FName)

        print("\n\n3. -------------DFT WATERMARK EXTRACTION --------")
        fw = glob.glob('.\watermarked\dft\*.png')
        NFiles = len(fw)
        print('Files In folder')
        print(fw)  # print all file names
        for Fc in range(0, NFiles):  # run till all fies processed
            stw = fw[Fc]  # select the file name
            FName = stw[18:len(stw)]
            FCoverImageName=FName[0:len(FName)-4]+'.jpg'
            coverImage = cv2.imread('./coverimages/'+FCoverImageName, 0)
            print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + stw)
            DFT(watermarkImage, FName)

        print("\n\n4. -------------SVD WATERMARK EXTRACTION --------")
        fw = glob.glob('.\watermarked\svd\*.png')
        NFiles = len(fw)
        print('Files In folder')
        print(fw)  # print all file names
        for Fc in range(0, NFiles):  # run till all fies processed
            stw = fw[Fc]  # select the file name
            FName = stw[18:len(stw)]
            FCoverImageName=FName[0:len(FName)-4]+'.jpg'
            coverImage = cv2.imread('./coverimages/'+FCoverImageName, 0)
            print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + stw)
            SVD(watermarkImage, FName)

        print("\n\n5. -------------DWT_SVD WATERMARK EXTRACTION --------")
        fw = glob.glob('.\watermarked\dwtsvd\*.png')
        NFiles = len(fw)
        print('Files In folder')
        print(fw)  # print all file names
        for Fc in range(0, NFiles):  # run till all fies processed
            stw = fw[Fc]  # select the file name
            FName = stw[21:len(stw)]
            FCoverImageName=FName[0:len(FName)-4]+'.jpg'
            coverImage = cv2.imread('./coverimages/'+FCoverImageName, 0)
            print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + stw)
            DWT_SVD(watermarkImage, FName)

        print("\n\n6. -------------DWT DCT SVD WATERMARK EXTRACTION --------")
        fw = glob.glob('.\watermarked\dwtdctsvd\*.png')
        NFiles = len(fw)
        print('Files In folder')
        print(fw)  # print all file names
        for Fc in range(0, NFiles):  # run till all fies processed
            stw = fw[Fc]  # select the file name
            FName = stw[24:len(stw)]
            FCoverImageName=FName[0:len(FName)-4]+'.jpg'
            coverImage = cv2.imread('./coverimages/'+FCoverImageName, 0)
            print('Processing File (' + str(Fc + 1) + '/' + str(NFiles) + '): ' + stw)
            DWT_DCT_SVD(watermarkImage, FName)
         
        exit(1)
